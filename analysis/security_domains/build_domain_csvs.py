#!/usr/bin/env python3
"""
Convert the sanitized Azure framework workbook + Word report into
per-domain CSVs that SecAI Radar can import, enriched with compliance
metadata pulled from the report.

Usage:
    python analysis/security_domains/build_domain_csvs.py
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook
import docx

# --------------------------------------------------------------------------------------
# Paths & constants

ROOT = Path(__file__).resolve().parents[2]  # repo root (SecAI)
RADAR_ROOT = ROOT / "secai-radar"

DEFAULT_XLSX = ROOT / "sanitized" / "Azure_Framework_2025_template_sanitized.xlsx"
DEFAULT_DOCX = ROOT / "sanitized" / "Cloud_Security_Assessment_Report_sanitized.docx"
DEFAULT_OUTPUT_DIR = ROOT / "analysis" / "security_domains" / "csv"
DEFAULT_MAPPING_JSON = ROOT / "analysis" / "security_domains" / "control_mappings.json"

DOMAIN_CODES_PATH = RADAR_ROOT / "seeds" / "domain_codes.json"

CONTROL_ROW_RE = re.compile(r"^([A-Z]{2,4}-\d+)\s*(.*)$")

# Column headers expected in the workbook
COL_CONTROL_ID = "Control ID"
COL_CONTROL_TITLE = "Control Title"
COL_SECURITY_PRINCIPLE = "Security Principle"
COL_AZURE_GUIDANCE = "Azure Guidance"
COL_MCSB = "MCSB Control ID Mapping"
COL_CIS = "CIS Mapping (v8)"
COL_NIST = "NIST SP 800-53 Mapping"
COL_PCI = "PCI-DSS Mapping (v3.2.1)"
COL_CSA = "CSA CCM Control ID"
COL_IANS = "IANS Azure Mapping"
COL_ASSESSMENT_QUESTION = "Assessment Question"
COL_EVIDENCE = "Evidence to Collect"
COL_EVIDENCE_COUNT = "Evidence Count"
COL_AVAILABLE_EVIDENCE_COUNT = "Available Evidence Count"
COL_RISK_SEVERITY = "Risk Severity (if non-compliant)"
COL_IMPL_GAP = "Implementation Gap"
COL_NOTES = "Notes/Comments"

# Output CSV schema (extras beyond API expectations are allowed)
CSV_HEADERS = [
    "ControlID",
    "Domain",
    "DomainCode",
    "ControlTitle",
    "ControlDescription",
    "Question",
    "RequiredEvidence",
    "EvidenceItems",
    "Status",
    "Owner",
    "Frequency",
    "ScoreNumeric",
    "Weight",
    "Notes",
    "SourceRef",
    "Tags",
    "UpdatedAt",
    "DocControlId",
    "ComplianceStatus",
    "Criticality",
    "Observations",
    "Gaps",
    "References",
    "EvidenceCount",
    "AvailableEvidenceCount",
    "EvidenceArtifacts",
    "EvidenceLastCollected",
    "ExcelControlId",
    "MCSBMapping",
    "CISMapping",
    "NISTMapping",
    "PCIDSSMapping",
    "CSACCMMapping",
    "IANSMapping",
    "RiskSeverity",
    "ImplementationGap",
]

CRITICALITY_TO_WEIGHT = {
    "critical": 1.0,
    "high": 0.85,
    "medium": 0.6,
    "low": 0.3,
}

STATUS_TO_SCORE = {
    "compliant": ( "Complete", 100 ),
    "partial compliant": ( "InProgress", 60 ),
    "partially compliant": ( "InProgress", 60 ),
    "non compliant": ( "NotStarted", 0 ),
    "not compliant": ( "NotStarted", 0 ),
    "not started": ( "NotStarted", 0 ),
    "in progress": ( "InProgress", 50 ),
}

SHEET_NAME_FIXUPS = {
    "posture and vulnerability manag": "Posture and vulnerability management",
}


# --------------------------------------------------------------------------------------
# Helpers

def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def normalize_domain_name(name: str) -> str:
    key = name.strip().lower()
    key = SHEET_NAME_FIXUPS.get(key, key)
    return key


def load_domain_codes() -> Dict[str, str]:
    data = json.loads(DOMAIN_CODES_PATH.read_text())
    return {k: v for k, v in data.items()}


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_section(rows: List[List[str]], label: str) -> str:
    label_lower = label.lower()
    capture = False
    collected: List[str] = []

    for row in rows:
        first = (row[0] if row else "").strip()
        if not capture:
            if first.lower().startswith(label_lower):
                capture = True
            continue
        if not first:
            if collected:
                break
            continue
        if first.endswith(":") and not first.lower().startswith(label_lower):
            break
        collected.append(first)

    return "\n".join(collected).strip()


def parse_doc_controls(doc_path: Path, domain_name_to_code: Dict[str, str]) -> Dict[str, dict]:
    document = docx.Document(doc_path)
    controls: Dict[str, dict] = {}

    for table_index, table in enumerate(document.tables):
        rows = [[cell.text.strip() for cell in row.cells] for row in table.rows]
        if not rows or len(rows) < 2:
            continue

        domain_candidate = rows[0][0].strip()
        domain_key = normalize_domain_name(domain_candidate)
        domain_code = domain_name_to_code.get(domain_key)
        if not domain_code:
            continue

        control_cell = rows[1][0].strip()
        match = CONTROL_ROW_RE.match(control_cell)
        if not match:
            continue

        doc_control_id = match.group(1).upper()
        control_title = match.group(2).strip()
        row1 = rows[1]

        compliance_status = ""
        criticality = ""
        for cell in row1:
            text = cell.strip()
            lower = text.lower()
            if lower.startswith("compliance status"):
                idx = row1.index(cell)
                if idx + 1 < len(row1):
                    compliance_status = row1[idx + 1].strip()
            if "criticality" in lower:
                criticality = text.split(":", 1)[-1].strip() if ":" in text else text

        observations = extract_section(rows, "Observations")
        gaps = extract_section(rows, "Gaps")
        references = extract_section(rows, "References")

        controls[doc_control_id] = {
            "doc_control_id": doc_control_id,
            "title": control_title,
            "domain": domain_candidate.strip(),
            "domainCode": domain_code,
            "compliance_status": compliance_status,
            "criticality": criticality,
            "observations": observations,
            "gaps": gaps,
            "references": references,
            "table_index": table_index,
        }

    return controls


def to_int(value) -> str:
    if value is None or value == "":
        return ""
    try:
        return str(int(float(value)))
    except (ValueError, TypeError):
        return str(value)


def split_evidence_items(text: str) -> str:
    text = clean(text)
    if not text:
        return ""
    items: List[str] = []
    for line in text.splitlines():
        cleaned = line.strip(" \t-•0123456789.")
        if cleaned:
            items.append(cleaned.strip())
    return " | ".join(items)


def build_description(row: dict) -> str:
    parts = []
    principle = clean(row.get(COL_SECURITY_PRINCIPLE, ""))
    guidance = clean(row.get(COL_AZURE_GUIDANCE, ""))
    if principle:
        parts.append(principle)
    if guidance:
        parts.append(guidance)
    return "\n\n".join(parts).strip()


def build_notes(row: dict, doc_record: Optional[dict]) -> str:
    parts: List[str] = []
    if doc_record and doc_record.get("observations"):
        parts.append(f"Observations: {doc_record['observations']}")
    if doc_record and doc_record.get("gaps"):
        parts.append(f"Gaps: {doc_record['gaps']}")
    impl_gap = clean(row.get(COL_IMPL_GAP, ""))
    if impl_gap:
        parts.append(f"Implementation Gap: {impl_gap}")
    notes = clean(row.get(COL_NOTES, ""))
    if notes:
        parts.append(notes)
    return "\n\n".join(parts).strip()


def build_tags(row: dict, doc_record: Optional[dict]) -> str:
    tags: List[str] = []
    for label, column in [
        ("CIS", COL_CIS),
        ("NIST", COL_NIST),
        ("PCI", COL_PCI),
        ("CSA", COL_CSA),
        ("IANS", COL_IANS),
        ("Risk", COL_RISK_SEVERITY),
    ]:
        value = clean(row.get(column))
        if value:
            tags.append(f"{label}:{value}")
    if doc_record and doc_record.get("compliance_status"):
        tags.append(f"DocStatus:{doc_record['compliance_status']}")
    return "; ".join(tags)


def build_source_ref(row: dict, doc_record: Optional[dict]) -> str:
    refs: List[str] = []
    excel_id = clean(row.get(COL_CONTROL_ID))
    if excel_id:
        refs.append(f"Excel={excel_id}")
    mcsb = clean(row.get(COL_MCSB))
    if mcsb:
        refs.append(f"MCSB={mcsb}")
    if doc_record:
        refs.append(f"DocID={doc_record['doc_control_id']}")
    cis = clean(row.get(COL_CIS))
    if cis:
        refs.append(f"CIS={cis}")
    nist = clean(row.get(COL_NIST))
    if nist:
        refs.append(f"NIST={nist}")
    return " | ".join(refs)


def determine_status_and_score(doc_record: Optional[dict]) -> tuple[str, int]:
    if not doc_record:
        return "NotStarted", 0
    status = doc_record.get("compliance_status", "").strip()
    lower = status.lower()
    if lower in STATUS_TO_SCORE:
        mapped_status, score = STATUS_TO_SCORE[lower]
        return mapped_status, score
    return "NotStarted", 0


def determine_weight(doc_record: Optional[dict]) -> float:
    if not doc_record:
        return 0.5
    criticality = doc_record.get("criticality", "").strip().lower()
    return CRITICALITY_TO_WEIGHT.get(criticality, 0.5)


def build_controls(
    workbook_path: Path,
    doc_controls: Dict[str, dict],
    domain_codes: Dict[str, str],
    tenant_id: str,
) -> List[dict]:
    wb = load_workbook(workbook_path, data_only=True)
    domain_name_to_code = {normalize_domain_name(v): k for k, v in domain_codes.items()}
    controls: List[dict] = []
    counters = defaultdict(int)
    missing_doc_refs: List[str] = []
    now_iso = datetime.now(timezone.utc).isoformat()

    for sheet in wb.sheetnames:
        if sheet in ("Introduction", "Summary", "Azure Landing Questions", "Setup-Questions", "Sources"):
            continue
        sheet_key = normalize_domain_name(sheet)
        domain_code = domain_name_to_code.get(sheet_key)
        if not domain_code:
            continue

        ws = wb[sheet]
        headers = []
        for idx, cell in enumerate(ws[1], start=1):
            header = cell.value
            if header is None:
                header = f"_col_{idx}"
            header = str(header).strip()
            if header == "":
                header = f"_col_{idx}"
            headers.append(header)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            excel_control_id = clean(row_dict.get(COL_CONTROL_ID))
            control_title = clean(row_dict.get(COL_CONTROL_TITLE))
            if not excel_control_id and not control_title:
                continue

            counters[domain_code] += 1
            control_id = f"SEC-{domain_code}-{counters[domain_code]:04d}"

            mcsb = clean(row_dict.get(COL_MCSB))
            doc_record = None
            if mcsb:
                doc_record = doc_controls.get(mcsb.upper())
                if not doc_record:
                    missing_doc_refs.append(f"{sheet}:{mcsb}")

            status, score = determine_status_and_score(doc_record)
            weight = determine_weight(doc_record)

            control_entry = {
                "ControlID": control_id,
                "Domain": domain_codes[domain_code],
                "DomainCode": domain_code,
                "ControlTitle": control_title,
                "ControlDescription": build_description(row_dict),
                "Question": clean(row_dict.get(COL_ASSESSMENT_QUESTION)),
                "RequiredEvidence": clean(row_dict.get(COL_EVIDENCE)),
                "EvidenceItems": split_evidence_items(row_dict.get(COL_EVIDENCE)),
                "Status": status,
                "Owner": "",
                "Frequency": "",
                "ScoreNumeric": score,
                "Weight": weight,
                "Notes": build_notes(row_dict, doc_record),
                "SourceRef": build_source_ref(row_dict, doc_record),
                "Tags": build_tags(row_dict, doc_record),
                "UpdatedAt": now_iso,
                "DocControlId": doc_record["doc_control_id"] if doc_record else "",
                "ComplianceStatus": doc_record.get("compliance_status") if doc_record else "",
                "Criticality": doc_record.get("criticality") if doc_record else "",
                "Observations": doc_record.get("observations") if doc_record else "",
                "Gaps": doc_record.get("gaps") if doc_record else "",
                "References": doc_record.get("references") if doc_record else "",
                "EvidenceCount": to_int(row_dict.get(COL_EVIDENCE_COUNT)),
                "AvailableEvidenceCount": to_int(row_dict.get(COL_AVAILABLE_EVIDENCE_COUNT)),
                "EvidenceArtifacts": "",
                "EvidenceLastCollected": "",
                "ExcelControlId": excel_control_id,
                "MCSBMapping": mcsb,
                "CISMapping": clean(row_dict.get(COL_CIS)),
                "NISTMapping": clean(row_dict.get(COL_NIST)),
                "PCIDSSMapping": clean(row_dict.get(COL_PCI)),
                "CSACCMMapping": clean(row_dict.get(COL_CSA)),
                "IANSMapping": clean(row_dict.get(COL_IANS)),
                "RiskSeverity": clean(row_dict.get(COL_RISK_SEVERITY)),
                "ImplementationGap": clean(row_dict.get(COL_IMPL_GAP)),
            }
            controls.append(control_entry)

    if missing_doc_refs:
        print(f"[WARN] Missing {len(missing_doc_refs)} doc mappings (see control_mappings.json for details).")
    return controls


def write_csv(path: Path, rows: Iterable[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fp:
        writer = csv.DictWriter(fp, fieldnames=CSV_HEADERS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in CSV_HEADERS})


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate per-domain CSVs for SecAI Radar.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Path to sanitized Azure framework workbook.")
    parser.add_argument("--docx", type=Path, default=DEFAULT_DOCX, help="Path to sanitized Word assessment report.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory for per-domain CSVs.")
    parser.add_argument("--mapping-json", type=Path, default=DEFAULT_MAPPING_JSON, help="Where to write the control mapping JSON.")
    parser.add_argument("--tenant", default="NICO", help="Tenant identifier (used for metadata only).")
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise FileNotFoundError(f"Workbook not found: {args.xlsx}")
    if not args.docx.exists():
        raise FileNotFoundError(f"Report not found: {args.docx}")

    domain_codes = load_domain_codes()
    name_to_code = {normalize_domain_name(v): k for k, v in domain_codes.items()}
    doc_controls = parse_doc_controls(args.docx, name_to_code)
    print(f"[INFO] Parsed {len(doc_controls)} controls from Word report.")

    controls = build_controls(args.xlsx, doc_controls, domain_codes, args.tenant)
    controls.sort(key=lambda r: r["ControlID"])

    # Write per-domain CSVs
    by_domain: Dict[str, List[dict]] = defaultdict(list)
    for row in controls:
        by_domain[row["DomainCode"]].append(row)

    for domain_code, rows in sorted(by_domain.items()):
        domain_name = domain_codes[domain_code]
        slug = slugify(domain_name)
        csv_path = args.output_dir / f"{domain_code}_{slug}.csv"
        write_csv(csv_path, rows)
        print(f"[INFO] Wrote {len(rows)} controls -> {csv_path}")

    # Aggregate CSV (useful for bulk import/testing)
    aggregate_path = args.output_dir / "ALL_CONTROLS.csv"
    write_csv(aggregate_path, controls)
    print(f"[INFO] Wrote aggregate CSV with {len(controls)} controls -> {aggregate_path}")

    # Mapping JSON (for evidence automation & UI linking)
    mapping = {
        row["ControlID"]: {
            "domain": row["Domain"],
            "domainCode": row["DomainCode"],
            "excelControlId": row["ExcelControlId"],
            "docControlId": row["DocControlId"],
            "mcsb": row["MCSBMapping"],
            "title": row["ControlTitle"],
            "status": row["Status"],
            "complianceStatus": row["ComplianceStatus"],
        }
        for row in controls
    }
    args.mapping_json.parent.mkdir(parents=True, exist_ok=True)
    args.mapping_json.write_text(json.dumps(mapping, indent=2))
    print(f"[INFO] Mapping JSON written -> {args.mapping_json}")


if __name__ == "__main__":
    main()
